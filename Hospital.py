import tkinter as tk
from tkinter import ttk
import sqlite3
import openpyxl
import os.path
import docx
import subprocess


class Main(tk.Frame):
    def __init__(self, root):
        super().__init__(root)
        self.init_main()
        self.db = db

    def init_main(self):
        toolbar = tk.Frame(bg='#ae8082', bd=8)
        toolbar.pack(side=tk.TOP, fill=tk.X)
        btn_ok = tk.Button(toolbar, text='Выбрать таблицу', bg='#caa9aa', bd=0, compound=tk.TOP)
        btn_ok.pack(side=tk.TOP)

        btn_ok.bind('<Button>', lambda event: self.GetTable(toolbar, combo_id.get()))

        combo_id = ttk.Combobox(toolbar)
        combo_id['values'] = ('Медицинские_Услуги', 'Оборудование', 'Пациенты', 'Сотрудники')
        combo_id.current(0)
        combo_id.pack(side=tk.TOP)

    def copy(self, NameTable):
        data = db.GetData(NameTable, True)
        my_wb = openpyxl.Workbook()
        my_sheet = my_wb.active

        for i in range(0, len(data['NAMES'])):
            cell = my_sheet.cell(row=1, column=i + 1)
            cell.value = data['NAMES'][i]

        for i in range(0, len(data['DATA'])):
            for j in range(0, len(data['DATA'][i])):
                cell = my_sheet.cell(row=i + 2, column=j + 1)
                cell.value = data['DATA'][i][j]

        my_wb.save(NameTable + ".xlsx")
        subprocess.call(NameTable + '.xlsx', shell=True)

    def copy_docx(self, NameTable):
        doc = docx.Document()
        data = db.GetData(NameTable, True)

        table = doc.add_table(rows=len(data['DATA']) + 1, cols=len(data['NAMES']))
        table.style = 'Table Grid'

        for i in range(0, len(data['NAMES'])):
            cell = table.cell(0, i)
            cell.text = data['NAMES'][i]

        for i in range(0, len(data['DATA'])):
            for j in range(0, len(data['DATA'][i])):
                cell = table.cell(i + 1, j)
                cell.text = str(data['DATA'][i][j])

        doc.save(NameTable + '.docx')
        subprocess.call(NameTable + '.docx', shell=True)

    def GetTable(self, toolbar, combo):
        global FlagTree
        global FlagBtn
        rows = self.db.GetNames(combo, True)

        if (FlagTree):
            self.tree.pack_forget()

        if (not FlagBtn):
            self.btn_add = tk.Button(toolbar, text='Добавить', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_add.pack(side=tk.LEFT)
            self.btn_add.bind('<Button>', lambda event: self.AddData(combo))
            self.btn_del = tk.Button(toolbar, text='Удалить', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_del.pack(side=tk.LEFT)
            self.btn_del.bind('<Button>', lambda event: self.DelData(combo))
            self.btn_red = tk.Button(toolbar, text='Редактировать', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_red.pack(side=tk.LEFT)
            self.btn_red.bind('<Button>', lambda event: self.EdData(combo))

            self.btn_copy = tk.Button(toolbar, text='Данные в электронную таблицу', bg='#caa9aa', bd=0,
                                      compound=tk.RIGHT)
            self.btn_copy.pack(side=tk.RIGHT)
            self.btn_copy.bind('<Button>', lambda event: self.copy(combo))
            self.btn_copy_docx = tk.Button(toolbar, text='Данные в текстовый редактор', bg='#caa9aa', bd=0,
                                           compound=tk.RIGHT)
            self.btn_copy_docx.pack(side=tk.RIGHT)
            self.btn_copy_docx.bind('<Button>', lambda event: self.copy_docx(combo))

            FlagBtn = True
        else:
            self.btn_add.pack_forget()
            self.btn_add = tk.Button(toolbar, text='Добавить', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_add.pack(side=tk.LEFT)
            self.btn_add.bind('<Button>', lambda event: self.AddData(combo))
            self.btn_del.pack_forget()
            self.btn_del = tk.Button(toolbar, text='Удалить', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_del.pack(side=tk.LEFT)
            self.btn_del.bind('<Button>', lambda event: self.DelData(combo))
            self.btn_red.pack_forget()
            self.btn_red = tk.Button(toolbar, text='Редактировать', bg='#caa9aa', bd=0, compound=tk.LEFT)
            self.btn_red.pack(side=tk.LEFT)
            self.btn_red.bind('<Button>', lambda event: self.EdData(combo))

            self.btn_copy.pack_forget()
            self.btn_copy = tk.Button(toolbar, text='Данные в эл.таблицу', bg='#caa9aa', bd=0, compound=tk.RIGHT)
            self.btn_copy.pack(side=tk.RIGHT)
            self.btn_copy.bind('<Button>', lambda event: self.copy(combo))
            self.btn_copy_docx.pack_forget()
            self.btn_copy_docx = tk.Button(toolbar, text='Данные в текстовый редактор', bg='#caa9aa', bd=0,
                                           compound=tk.RIGHT)
            self.btn_copy_docx.pack(side=tk.RIGHT)
            self.btn_copy_docx.bind('<Button>', lambda event: self.copy_docx(combo))

        self.tree = ttk.Treeview(self, columns=(rows), height=20, show='headings')
        width_rows = 0
        for i in range(0, len(rows)):
            width_rows += len(rows[i]) * 20
            self.tree.column(rows[i], width=len(rows[i]) * 20, anchor=tk.CENTER)
            self.tree.heading(rows[i], text=rows[i])
        width_rows += 40
        self.tree.pack()
        FlagTree = True
        root.geometry(str(width_rows) + "x600+300+100")

        self.insert_data(combo)

    def insert_data(self, combo):
        data = self.db.GetData(combo, True)
        # Пузырёк:
        for i in range(1, len(data['DATA'])):
            if (data['DATA'][i - 1][0] > data['DATA'][i][0]):
                data['DATA'][i - 1][0], data['DATA'][i][0] = data['DATA'][i][0], data['DATA'][i - 1][0]
        # Конец пузырька
        [self.tree.delete(i) for i in self.tree.get_children()]
        [self.tree.insert('', 'end', values=row) for row in data['DATA']]

    def AddData(self, NameTable):
        Child("add", NameTable)

    def DelData(self, NameTable):
        Child("del", NameTable)

    def EdData(self, NameTable):
        Child("red", NameTable)


class Child(tk.Toplevel):
    def __init__(self, TypeView, NameTable):
        super().__init__(root)
        self.init_child(TypeView, NameTable)
        self.view = app
        self.db = db

    def init_child(self, TypeView, NameTable):
        self.resizable(False, False)
        self.grab_set()
        self.focus_set()

        if (TypeView == 'add'):

            NamesRow = db.GetNames(NameTable, True)
            LenRow = len(NamesRow) * 30 + 100

            self.geometry('500x' + str(LenRow) + '+350+150')
            self.title('Добавить')

            entry = []

            for i in range(1, len(NamesRow)):
                label = ttk.Label(self, text=NamesRow[i])
                label.place(x=50, y=i * 30)

                entry.append(ttk.Entry(self))
                entry[i - 1].place(x=200, y=i * 30)

            self.label_error = ttk.Label(self, text='')

            btn_ok = ttk.Button(self, text='OK')
            btn_ok.place(x=400, y=90)
            btn_ok.bind('<Button>', lambda event: self.check(self, entry, NameTable, i * 30 + 30, False))

        elif (TypeView == 'del' or TypeView == 'red'):
            self.title('Удалить')
            if (TypeView == 'red'):
                self.title('Редактировать')
            self.geometry('500x90+350+150')

            self.label = ttk.Label(self, text='ID записи')
            self.label.place(x=50, y=30)

            self.entry = ttk.Entry(self)
            self.entry.place(x=200, y=30)

            self.label_error = ttk.Label(self, text='')

            self.btn_ok = ttk.Button(self, text='OK')
            self.btn_ok.place(x=400, y=30)
            self.btn_ok.bind('<Button-1>', lambda event: self.check_id(self, self.entry.get(), NameTable, TypeView))

    def check_id(self, this, entry, NameTable, TypeView):
        data = self.db.GetData(NameTable, True)

        if (entry != ''):
            try:
                int(entry)
            except ValueError:
                self.label_error.grid(row=1, column=1)
                self.label_error['text'] = "Ошибка типа данных ID: введена строка"
                self.label_error.place(x=50, y=60)
            else:
                count = True
                for i in range(0, len(data['DATA'])):
                    if (int(data['DATA'][i][0]) == int(entry)):
                        count = False
                        if (TypeView == 'del'):
                            self.db.delete_data(NameTable, entry)
                            self.view.insert_data(NameTable)
                            this.destroy()
                        else:
                            self.btn_ok.destroy()
                            self.label_error.destroy()
                            self.entry.destroy()
                            self.label.destroy()
                            self.editing(this, data['DATA'][i], data['NAMES'], NameTable, entry)
                if (count):
                    self.label_error.grid(row=1, column=1)
                    self.label_error['text'] = "Ошибка: не существует данного ID"
                    self.label_error.place(x=50, y=60)

    def editing(self, this, data, names, NameTable, RecID):

        LenRow = len(names) * 30 + 100

        self.geometry('500x' + str(LenRow) + '+350+150')

        entry = []

        for i in range(1, len(names)):
            label = ttk.Label(self, text=names[i])
            label.place(x=50, y=i * 30)

            entry.append(ttk.Entry(self))
            entry[i - 1].place(x=200, y=i * 30)
            entry[i - 1].insert(0, data[i])

        self.label_error = ttk.Label(self, text='')

        btn_ok = ttk.Button(self, text='OK')
        btn_ok.place(x=390, y=90)
        btn_ok.bind('<Button>', lambda event: self.check(self, entry, NameTable, i * 30 + 30, True, RecID))

    def check(self, this, entry, NameTable, Y, flagRed, RecID=0):
        NamesRow = db.GetNames(NameTable, False)

        resType = []
        for i in range(1, len(NamesRow)):
            resType.append([NamesRow[i][2], NamesRow[i][1]])

        entry_val = {}

        for i in range(0, len(entry)):
            if (entry[i].get() != '' and resType[i][0] == 'integer'):
                try:
                    int(entry[i].get())
                except ValueError:
                    self.label_error.grid(row=1, column=1)
                    self.label_error['text'] = "Ошибка типа данных " + resType[i][1]
                    self.label_error.place(x=50, y=Y)
                else:
                    entry_val[resType[i][1]] = entry[i].get()
                    if (len(entry_val) == len(entry)):
                        if (flagRed):
                            self.db.edit_data(NameTable, entry_val, RecID)
                        else:
                            self.db.insert_data(NameTable, entry_val)
                        self.view.insert_data(NameTable)
                        this.destroy()
            elif (entry[i].get() != '' and resType[i][0] == 'real'):
                try:
                    float(entry[i].get())
                except ValueError:
                    self.label_error.grid(row=1, column=1)
                    self.label_error['text'] = "Ошибка типа данных " + resType[i][1]
                    self.label_error.place(x=50, y=Y)
                else:
                    entry_val[resType[i][1]] = entry[i].get()
                    if (len(entry_val) == len(entry)):
                        if (flagRed):
                            self.db.edit_data(NameTable, entry_val, RecID)
                        else:
                            self.db.insert_data(NameTable, entry_val)
                        self.view.insert_data(NameTable)
                        this.destroy()
            else:
                entry_val[resType[i][1]] = entry[i].get()
                if (len(entry_val) == len(entry)):
                    if (flagRed):
                        self.db.edit_data(NameTable, entry_val, RecID)
                    else:
                        self.db.insert_data(NameTable, entry_val)
                    self.view.insert_data(NameTable)
                    this.destroy()


class DB:
    def __init__(self):
        self.conn = sqlite3.connect('Medical_services.db')
        self.c = self.conn.cursor()
        self.c.execute('''CREATE TABLE IF NOT EXISTS Медицинские_Услуги (
    ID              INTEGER PRIMARY KEY AUTOINCREMENT,
    ID_Пациента     INTEGER NOT NULL
                            REFERENCES Пациенты (ID),
    ВидУслуги       TEXT    NOT NULL,
    СтоимостьУслуги REAL    NOT NULL)''')

        self.c.execute('''CREATE TABLE IF NOT EXISTS Оборудование (
    ID                   INTEGER PRIMARY KEY AUTOINCREMENT,
    НазваниеОборудования TEXT    NOT NULL,
    Количество           INTEGER NOT NULL,
    Стоимость            REAL    NOT NULL)''')
        self.c.execute('''CREATE TABLE IF NOT EXISTS Пациенты (
    ID            INTEGER PRIMARY KEY AUTOINCREMENT,
    ID_Сотрудника INTEGER NOT NULL
                          REFERENCES Сотрудники,
    Фамилия       TEXT    NOT NULL,
    Имя           TEXT    NOT NULL,
    Отчество      TEXT    NOT NULL,
    ДатаРождения  DATE    NOT NULL)''')
        self.c.execute('''CREATE TABLE IF NOT EXISTS Сотрудники (
    ID              INTEGER PRIMARY KEY AUTOINCREMENT,
    ID_Оборудования INTEGER NOT NULL
                            REFERENCES Оборудование (ID),
    ID_Услуги       INTEGER NOT NULL
                            REFERENCES Медицинские_Услуги (ID),
    Фамилия         TEXT    NOT NULL,
    Имя             TEXT    NOT NULL,
    Отчество        TEXT    NOT NULL,
    Должность       TEXT    NOT NULL)''')
        self.conn.commit()

    def insert_data(self, table, values):
        keys = ''
        vals = []
        lan = ''
        for k, v in values.items():
            keys += k + ', '
            vals.append(v)
            lan += '?, '
        self.c.execute('''INSERT INTO ''' + table + '''(''' + keys[0:-2] + ''') VALUES (''' + lan[0:-2] + ''')''', vals)
        self.conn.commit()

    def edit_data(self, table, values, RecID):
        self.c.execute('''DELETE FROM ''' + table + ''' WHERE ID = ''' + RecID)
        keys = ''
        vals = [RecID]
        lan = ''
        for k, v in values.items():
            keys += k + ', '
            vals.append(v)
            lan += '?, '
        print(vals)
        self.c.execute(
            '''INSERT INTO ''' + table + '''(ID, ''' + keys[0:-2] + ''') VALUES (?, ''' + lan[0:-2] + ''')''', vals)
        self.conn.commit()

    def delete_data(self, table, value):
        self.c.execute('''DELETE FROM ''' + table + ''' WHERE ID = ''' + value)
        self.conn.commit()

    def GetData(self, combo, flag):
        self.c.execute('''PRAGMA table_info(''' + combo + ''')''')
        names = self.c.fetchall()
        if (flag):
            rows = []
            for i in range(0, len(names)):
                rows.append(names[i][1])
        self.c.execute('''SELECT * FROM ''' + combo + ''' ORDER BY ID ASC''')
        data = self.c.fetchall()
        self.conn.commit()
        result = {"NAMES": rows, "DATA": data}
        return result

    def GetNames(self, combo, flag):
        self.c.execute('''PRAGMA table_info(''' + combo + ''')''')
        names = self.c.fetchall()
        if (flag):
            result = []
            for i in range(0, len(names)):
                result.append(names[i][1])
            return result
        return names


if __name__ == "__main__":
    FlagTree = False
    FlagBtn = False
    root = tk.Tk()
    db = DB()
    app = Main(root)
    app.pack()
    root.title("Медицинские_Услуги.Инвентаризация_оборудования")
    root.geometry("800x700+300+100")
    root.mainloop()
